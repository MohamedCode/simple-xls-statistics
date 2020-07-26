import openpyxl as xl


def get_total_working_hrs_by_project(file_path):
    wb = xl.load_workbook(file_path)
    sheet_names = wb.get_sheet_names()
    total_hours = []

    for sh in sheet_names:
        sheet = wb[sh]
        for row in range(2, sheet.max_row):
            cell = sheet.cell(row, 9)
            if cell.value != None \
                and (isinstance(cell.value, float)
                     or isinstance(cell.value, int)
            ):
                total_hours.append(cell.value)

    return sum(total_hours)


def get_sheet_sum(file_path):
    wb = xl.load_workbook(file_path)
    sheet_names = wb.get_sheet_names()
    return len(sheet_names)
